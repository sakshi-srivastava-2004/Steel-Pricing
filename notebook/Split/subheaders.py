import openpyxl
import pandas as pd
import os

# === CONFIGURATION ===
INPUT_FILE = 'melting scrap_D_dom.xlsx'
YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}
OUTPUT_DIR = 'melting scrap_D_dom'

os.makedirs(OUTPUT_DIR, exist_ok=True)  # Create output directory if not exist

def is_yellow(cell):
    fill = cell.fill
    if fill.fill_type == 'solid':
        color = fill.start_color
        if color.rgb:
            return color.rgb.upper() in YELLOW_RGB_VALUES
        if color.indexed == 6:
            return True
    return False

def find_yellow_header_row(ws, max_rows=10):
    for row_num in range(1, max_rows + 1):
        if any(is_yellow(cell) for cell in ws[row_num]):
            return row_num
    return None

def get_merged_range_for_cell(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

def sanitize_filename(name):
    return name.strip().replace('/', '_').replace('\\', '_').replace(':', '-').replace('?', '').replace('*', '')[:50]

def main():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing Sheet: {sheet_name}")

        header_row_num = find_yellow_header_row(ws)
        if not header_row_num:
            print(f"No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
            continue

        subheader_row_num = header_row_num + 1
        df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, header=subheader_row_num - 1)
        df.columns = [str(col).strip() for col in df.columns]

        header_row = ws[header_row_num]

        week_col_idx = None
        week_col_name = None

        # Try to find a column named 'week' to include optionally
        for cell in header_row:
            if cell.value and 'week' in str(cell.value).strip().lower():
                week_col_idx = cell.col_idx
                week_col_name = ws.cell(row=subheader_row_num, column=week_col_idx).value
                if week_col_name:
                    week_col_name = str(week_col_name).strip()
                print(f"→ Found 'week' column: {week_col_name}")
                break

        processed_headers = set()

        for cell in header_row:
            if is_yellow(cell):
                yellow_header = str(cell.value).strip()
                if yellow_header in processed_headers:
                    continue
                processed_headers.add(yellow_header)

                merged_range = get_merged_range_for_cell(ws, cell)
                min_col = merged_range.min_col if merged_range else cell.col_idx
                max_col = merged_range.max_col if merged_range else cell.col_idx

                print(f"\nExtracting group under yellow header: '{yellow_header}' spanning cols {min_col} to {max_col}")

                subheaders = []
                for col_idx in range(min_col, max_col + 1):
                    sub_cell = ws.cell(row=subheader_row_num, column=col_idx)
                    sub_value = sub_cell.value
                    if sub_value:
                        sub_value_cleaned = str(sub_value).strip()
                        if sub_value_cleaned in df.columns:
                            subheaders.append(sub_value_cleaned)

                if week_col_name and week_col_name in df.columns and week_col_name not in subheaders:
                    subheaders.insert(0, week_col_name)

                if not subheaders:
                    print(f"→ No valid subheaders found for '{yellow_header}'. Skipping.")
                    continue

                extracted_df = df[subheaders].dropna(how='all')  # Drop fully empty rows

                # Save to a separate file
                safe_name = sanitize_filename(f"{sheet_name}_{yellow_header}")
                output_path = os.path.join(OUTPUT_DIR, f"{safe_name}.xlsx")
                extracted_df.to_excel(output_path, index=False)
                print(f"✔ Saved to '{output_path}' ({len(extracted_df)} rows, {len(extracted_df.columns)} columns)")

if __name__ == "__main__":
    main()
