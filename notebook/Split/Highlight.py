import openpyxl
import pandas as pd
import re
import numpy as np
import os
import glob
from datetime import datetime, timedelta

# --- Configuration ---
# IMPORTANT: Change this to the actual directory path where your Excel files are located
input_directory = 'C:/intership/notebook/input'
# The name of the the single consolidated output Excel file
consolidated_output_file_name = os.path.join(input_directory, 'consolidated_processed_data.xlsx')

# RGB values for yellow color used in Excel for header detection
YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'} # Add other common yellow RGBs if needed

# --- Utility Functions ---

def is_yellow(cell):
    """Checks if a cell's fill color is yellow."""
    fill = cell.fill
    if fill.fill_type == 'solid':
        color = fill.start_color
        if color.rgb:
            return color.rgb.upper() in YELLOW_RGB_VALUES
        if color.indexed is not None:
            # Indexed color 6 is often yellow in default palettes
            return color.indexed == 6
    return False

def find_yellow_header_row(ws, max_rows=10):
    """Finds the row number of the first row containing any yellow-highlighted cell."""
    for row_num in range(1, max_rows + 1):
        row = ws[row_num]
        if any(is_yellow(cell) for cell in row):
            return row_num
    return None

def get_merged_range_for_cell(ws, cell):
    """
    Returns the merged cell range object if the given cell is part of one,
    otherwise returns None.
    """
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

def find_nearest_non_h(series, index):
    """
    Finds the nearest non-'H' numeric value in a pandas Series.
    It looks backward first, then forward.
    """
    # Look backward
    for i in range(index - 1, -1, -1):
        val = series.iloc[i]
        if pd.notna(val) and val != 'H':
            try:
                return float(val)
            except ValueError:
                pass
    # Look forward
    for i in range(index + 1, len(series)):
        val = series.iloc[i]
        if pd.notna(val) and val != 'H':
            try:
                return float(val)
            except ValueError:
                pass
    return np.nan # If no nearest numeric value found


def calculate_weekly_averages_with_regions(df_input, date_col_name, material_col_name, region_col_name, price_col_name):
    """
    Processes the DataFrame to calculate weekly averages, handles 'H' values,
    and associates data with material and region.
    Assumes df_input contains 'Date', 'Material', 'Region', and 'Price' columns.
    """
    df = df_input.copy()

    # Drop rows where material, region, or price are entirely missing
    df = df.dropna(subset=[material_col_name, region_col_name, price_col_name], how='all')

    # Convert 'Date' column to datetime objects
    df[date_col_name] = pd.to_datetime(df[date_col_name], errors='coerce')
    df = df.dropna(subset=[date_col_name]) # Drop rows where date conversion failed

    if df.empty:
        return pd.DataFrame()

    # Sort by date to ensure correct 'H' replacement and week calculation
    df = df.sort_values(by=date_col_name).reset_index(drop=True)

    # Convert price column to numeric, coercing errors
    # Replace common empty/placeholder strings with NaN for proper numeric conversion
    df[price_col_name] = df[price_col_name].replace(['-', '', ' '], np.nan)
    
    # Iterate through each material and region group to replace 'H' values
    processed_groups = []
    for (material_val, region_val), group_df in df.groupby([material_col_name, region_col_name], dropna=False):
        
        # Replace 'H' values in the price column
        price_series = group_df[price_col_name].astype(str).str.strip().str.upper()
        h_indices = price_series[price_series == 'H'].index # Get original indices
        
        for original_idx in h_indices:
            # Find the nearest non-H value from the group's price column
            # Need to get the index within the *group_df* for find_nearest_non_h
            relative_idx_in_group = group_df.index.get_loc(original_idx)
            nearest_val = find_nearest_non_h(group_df[price_col_name], relative_idx_in_group)
            
            # Update the original DataFrame using .loc and the original index
            df.loc[original_idx, price_col_name] = nearest_val
            
        processed_groups.append(group_df)
    
    # If there were groups, reconstruct the DataFrame, otherwise use the original for type conversion
    if processed_groups:
        df = pd.concat(processed_groups).sort_index() # Reconstruct and sort by original index
    
    # Ensure price column is numeric after replacement
    df[price_col_name] = pd.to_numeric(df[price_col_name], errors='coerce')

    # Group by weeks starting Monday (weekday 0) to Sunday (weekday 6)
    # Calculate the start of the week (Monday) for each date
    df['Week_Start_Date'] = df[date_col_name] - pd.to_timedelta(df[date_col_name].dt.dayofweek, unit='D')

    all_weekly_data = []
    
    # Determine the global start date for calculating week numbers
    min_date = df[date_col_name].min()
    global_week_start_reference = min_date - timedelta(days=min_date.weekday()) # This will be the first Monday or earlier

    # Group by material, region, and then by the calculated week start date
    for (material_val, region_val, week_start_date), group in df.groupby([material_col_name, region_col_name, 'Week_Start_Date'], dropna=False):
        if pd.isna(material_val) or pd.isna(region_val) or group.empty:
            continue

        # Calculate the sequential week number
        week_number = int((week_start_date - global_week_start_reference).days / 7) + 1

        avg_price = group[price_col_name].mean()

        all_weekly_data.append({
            'Week': week_number,
            'material': material_val,
            'region': region_val,
            'price': avg_price
        })

    result_df = pd.DataFrame(all_weekly_data)
    # Drop rows where price is NaN (e.g., if a whole week had no valid numeric data after H replacement)
    result_df.dropna(subset=['price'], inplace=True)
    return result_df

# --- Main Processing Logic ---

def main():
    excel_files = glob.glob(os.path.join(input_directory, '*.xlsx'))

    if not excel_files:
        print(f"No Excel files found in the directory: {input_directory}")
        return

    print(f"Found {len(excel_files)} Excel files to process in {input_directory}:")
    for file_path in excel_files:
        print(f"- {os.path.basename(file_path)}")

    final_processed_data_frames = [] # To store DataFrames ready for final consolidation

    for file_path in excel_files:
        print(f"\n--- Processing file: {os.path.basename(file_path)} ---")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            print(f"An error occurred while opening '{os.path.basename(file_path)}': {e}. Skipping.")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Processing Sheet: {sheet_name}")

            header_row_num = find_yellow_header_row(ws)
            if not header_row_num:
                print(f"  No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
                continue

            subheader_row_num = header_row_num + 1
            max_row = ws.max_row
            max_col = ws.max_column

            # Identify the global Date column index
            global_date_col_idx = None
            for col in range(1, max_col + 1):
                header_val = str(ws.cell(row=header_row_num, column=col).value).strip().lower()
                subheader_val = str(ws.cell(row=subheader_row_num, column=col).value).strip().lower()
                if 'date' in header_val or 'date' in subheader_val:
                    global_date_col_idx = col
                    break
            
            if not global_date_col_idx:
                print(f"  No 'Date' column found in sheet '{sheet_name}'. Cannot calculate weekly averages. Skipping.")
                continue

            # Iterate through columns to find material blocks (yellow headers)
            processed_materials = set() # To prevent re-processing merged yellow headers
            for col_idx in range(1, max_col + 1):
                header_cell = ws.cell(row=header_row_num, column=col_idx)

                if is_yellow(header_cell) and header_cell.value:
                    material_name = str(header_cell.value).strip()

                    # If this material has already been processed due to a merged cell, skip
                    if material_name in processed_materials:
                        continue
                    processed_materials.add(material_name) # Mark as processed

                    # Determine the column range for the current material block
                    merged_range = get_merged_range_for_cell(ws, header_cell)
                    min_material_col = merged_range.min_col if merged_range else header_cell.col_idx
                    max_material_col = merged_range.max_col if merged_range else header_cell.col_idx

                    # Collect region-price column pairs under this material
                    region_price_col_map = {} # {region_name: col_index}
                    for current_col_idx_in_block in range(min_material_col, max_material_col + 1):
                        subheader_cell = ws.cell(row=subheader_row_num, column=current_col_idx_in_block)
                        region_val = str(subheader_cell.value).strip() if subheader_cell.value is not None else ""
                        
                        # Assuming direct subheaders under yellow material are regions
                        if region_val and current_col_idx_in_block != global_date_col_idx:
                            # Avoid picking up 'date' as a region if it's within the block
                            region_price_col_map[region_val] = current_col_idx_in_block

                    if not region_price_col_map:
                        print(f"    No specific region/price columns found under material '{material_name}'. Skipping.")
                        continue

                    # Extract data for the current material block for all identified regions
                    data_for_material = []
                    for r_idx in range(subheader_row_num + 1, max_row + 1):
                        date_val = ws.cell(row=r_idx, column=global_date_col_idx).value
                        
                        # Stop processing rows if the date column is empty
                        if date_val is None or str(date_val).strip() == '':
                            break
                        
                        for region, price_col_idx in region_price_col_map.items():
                            price_val = ws.cell(row=r_idx, column=price_col_idx).value
                            data_for_material.append({
                                'Date': date_val,
                                'Material_Name': material_name, # Use a temporary name to avoid conflict with desired 'material' output
                                'Region_Name': region,         # Use a temporary name for region
                                'Price_Value': price_val       # Use a temporary name for price
                            })

                    if not data_for_material:
                        print(f"    No data found for material '{material_name}'. Skipping.")
                        continue

                    df_raw_material_data = pd.DataFrame(data_for_material)

                    # Now, call the weekly averaging function
                    processed_material_df = calculate_weekly_averages_with_regions(
                        df_raw_material_data,
                        date_col_name='Date',
                        material_col_name='Material_Name',
                        region_col_name='Region_Name',
                        price_col_name='Price_Value'
                    )

                    if not processed_material_df.empty:
                        final_processed_data_frames.append(processed_material_df)
                    else:
                        print(f"    No valid weekly averaged data generated for material '{material_name}'.")

    # --- Final Step: Write all consolidated data to a single Excel file ---
    if final_processed_data_frames:
        final_consolidated_df = pd.concat(final_processed_data_frames, ignore_index=True)
        
        # Drop duplicates based on the final output columns
        final_consolidated_df.drop_duplicates(subset=['Week', 'material', 'region', 'price'], inplace=True)
        
        # Sort by Week, material, and region for a clean output
        final_consolidated_df.sort_values(by=['Week', 'material', 'region'], inplace=True)

        try:
            with pd.ExcelWriter(consolidated_output_file_name, engine='xlsxwriter') as writer:
                # Write the DataFrame to a sheet named 'Consolidated Data'
                final_consolidated_df.to_excel(writer, sheet_name='Consolidated Data', index=False)
            print(f"\n✅ All consolidated processed data saved to '{consolidated_output_file_name}'")
        except Exception as e:
            print(f"An error occurred while writing the consolidated Excel file: {e}")
    else:
        print(f"\nNo data was processed from any file to create '{consolidated_output_file_name}'.")

if __name__ == "__main__":
    main()