# import openpyxl
# import pandas as pd
# import numpy as np

# # === CONFIGURATION ===
# INPUT_FILE = 'billet_D_dom.xlsx'  # Your input Excel file (with yellow headers)
# OUTPUT_FILE = 'billet_averaged.xlsx'
# YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

# # === HELPER FUNCTIONS ===
# def is_yellow(cell):
#     fill = cell.fill
#     if fill.fill_type == 'solid':
#         color = fill.start_color
#         if color.rgb:
#             return color.rgb.upper() in YELLOW_RGB_VALUES
#         if color.indexed == 6:
#             return True
#     return False

# def find_yellow_header_row(ws, max_check=10):
#     for r in range(1, max_check + 1):
#         if any(is_yellow(cell) for cell in ws[r]):
#             return r
#     return None

# def get_merged_range_for_cell(ws, cell):
#     for mr in ws.merged_cells.ranges:
#         if cell.coordinate in mr:
#             return mr
#     return None

# def average_weeks_only(df: pd.DataFrame) -> pd.DataFrame:
#     df = df.replace(['-', '', ' '], np.nan).reset_index(drop=True)
#     weeks = []
#     start = 0
#     week_num = 1

#     col0 = df.iloc[:, 0].astype(str).str.strip().str.upper()
#     for i, val in enumerate(col0):
#         if val == 'H':
#             block = df.loc[start:i]
#             if not block.empty:
#                 num_avg = block.select_dtypes(include=[np.number]).mean()
#                 cat_avg = block.select_dtypes(exclude=[np.number]).apply(
#                     lambda c: c.dropna().iloc[0] if c.dropna().any() else np.nan)
#                 row = pd.concat([pd.Series([week_num], index=['Week_Number']), cat_avg, num_avg])
#                 weeks.append(row)
#                 week_num += 1
#             start = i + 1

#     tail = df.loc[start:]
#     if not tail.empty:
#         num_avg = tail.select_dtypes(include=[np.number]).mean()
#         cat_avg = tail.select_dtypes(exclude=[np.number]).apply(
#             lambda c: c.dropna().iloc[0] if c.dropna().any() else np.nan)
#         row = pd.concat([pd.Series([week_num], index=['Week_Number']), cat_avg, num_avg])
#         weeks.append(row)

#     return pd.DataFrame(weeks)

# # === MAIN PROCESS ===
# wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
# writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')

# for sheet_name in wb.sheetnames:
#     ws = wb[sheet_name]
#     hr = find_yellow_header_row(ws)
#     if hr is None:
#         continue

#     sub_hr = hr + 1
#     df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, header=sub_hr - 1)
#     df.columns = [str(c).strip() for c in df.columns]

#     for cell in ws[hr]:
#         if not is_yellow(cell):
#             continue

#         mr = get_merged_range_for_cell(ws, cell)
#         c1 = mr.min_col if mr else cell.col_idx
#         c2 = mr.max_col if mr else c1

#         subs = []
#         for c in range(c1, c2 + 1):
#             val = ws.cell(row=sub_hr, column=c).value
#             if val and str(val).strip() in df.columns:
#                 subs.append(str(val).strip())

#         if not subs:
#             continue

#         subdf = df[subs]
#         avgdf = average_weeks_only(subdf)

#         name = str(cell.value).strip()[:30].replace('/', '_').replace('\\', '_')
#         out_name = f"{sheet_name}_{name}"
#         avgdf.to_excel(writer, sheet_name=out_name, index=False)

# writer.close()
# print(f" All groups processed. Output saved as '{OUTPUT_FILE}'.")


import openpyxl
import pandas as pd
import numpy as np

# === CONFIGURATION ===
INPUT_FILE = 'ingot_D_dom.xlsx'  # Your input Excel file (with yellow headers)
OUTPUT_FILE = 'ingot_averaged.xlsx'
YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

# === HELPER FUNCTIONS ===
def is_yellow(cell):
    fill = cell.fill
    if fill.fill_type == 'solid':
        color = fill.start_color
        if color.rgb:
            return color.rgb.upper() in YELLOW_RGB_VALUES
        if color.indexed == 6:
            return True
    return False

def find_yellow_header_row(ws, max_check=10):
    for r in range(1, max_check + 1):
        if any(is_yellow(cell) for cell in ws[r]):
            return r
    return None

def get_merged_range_for_cell(ws, cell):
    for mr in ws.merged_cells.ranges:
        if cell.coordinate in mr:
            return mr
    return None

def average_weeks_only(df: pd.DataFrame) -> pd.DataFrame:
    df = df.replace(['-', '', ' '], np.nan).reset_index(drop=True)
    weeks = []
    start = 0
    week_num = 1

    col0 = df.iloc[:, 0].astype(str).str.strip().str.upper()
    h_indices = col0[col0 == 'H'].index.tolist()

    # Add the end of the DataFrame as the final boundary
    boundaries = h_indices + [len(df)]

    for end in boundaries:
        block = df.iloc[start:end].copy()
        # Exclude rows with 'H' in first column from averaging block
        block = block[~block.iloc[:, 0].astype(str).str.upper().eq('H')]

        if not block.empty:
            num_avg = block.select_dtypes(include=[np.number]).mean()
            cat_avg = block.select_dtypes(exclude=[np.number]).apply(
                lambda c: c.dropna().iloc[0] if c.dropna().any() else np.nan
            )
            row = pd.concat([pd.Series([week_num], index=['Week_Number']), cat_avg, num_avg])
            weeks.append(row)
            week_num += 1

        start = end + 1

    return pd.DataFrame(weeks)

# === MAIN PROCESS ===
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    hr = find_yellow_header_row(ws)
    if hr is None:
        continue

    sub_hr = hr + 1
    df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, header=sub_hr - 1)
    df.columns = [str(c).strip() for c in df.columns]

    for cell in ws[hr]:
        if not is_yellow(cell):
            continue

        mr = get_merged_range_for_cell(ws, cell)
        c1 = mr.min_col if mr else cell.col_idx
        c2 = mr.max_col if mr else c1

        subs = []
        for c in range(c1, c2 + 1):
            val = ws.cell(row=sub_hr, column=c).value
            if val and str(val).strip() in df.columns:
                subs.append(str(val).strip())

        if not subs:
            continue

        subdf = df[subs]
        avgdf = average_weeks_only(subdf)

        name = str(cell.value).strip()[:30].replace('/', '_').replace('\\', '_')
        out_name = f"{sheet_name}_{name}"
        avgdf.to_excel(writer, sheet_name=out_name, index=False)

writer.close()
print(f"All groups processed. Output saved as '{OUTPUT_FILE}'.")
