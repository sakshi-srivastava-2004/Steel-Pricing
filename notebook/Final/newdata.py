# import openpyxl

# file_path = 'rebar_dom.xlsx'  # Replace this

# YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

# def is_yellow(cell):
#     fill = cell.fill
#     if fill.fill_type == 'solid':
#         color = fill.start_color
#         if color.rgb:
#             return color.rgb.upper() in YELLOW_RGB_VALUES
#         if color.indexed is not None:
#             return color.indexed == 6
#     return False

# def find_yellow_header_row(ws, max_rows=10):
#     for row_num in range(1, max_rows+1):
#         row = ws[row_num]
#         if any(is_yellow(cell) for cell in row):
#             return row_num
#     return None

# def get_merged_range_for_cell(ws, cell):
#     for merged_range in ws.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             return merged_range
#     return None

# def main():
#     wb = openpyxl.load_workbook(file_path, data_only=True)
#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         print(f"\nSheet: {sheet_name}")

#         header_row_num = find_yellow_header_row(ws, max_rows=10)
#         if not header_row_num:
#             print("No yellow highlighted header row found.")
#             continue

#         header_row = ws[header_row_num]

#         # The row number of the subheader (row below the yellow header)
#         subheader_row_num = header_row_num + 1
#         subheader_row = ws[subheader_row_num]

#         # For each yellow cell, find the columns it covers
#         for cell in header_row:
#             if is_yellow(cell):
#                 merged_range = get_merged_range_for_cell(ws, cell)
#                 if merged_range:
#                     # Columns spanned by the merged cell
#                     min_col = merged_range.min_col
#                     max_col = merged_range.max_col
#                 else:
#                     min_col = cell.col_idx
#                     max_col = cell.col_idx

#                 yellow_header = cell.value
#                 print(f"\nYellow header: '{yellow_header}' spans columns {min_col} to {max_col}")

#                 # Extract subheaders in the row below for these columns
#                 subheaders = []
#                 for col in range(min_col, max_col + 1):
#                     sub_cell = ws.cell(row=subheader_row_num, column=col)
#                     subheaders.append(sub_cell.value)

#                 print(f"Subheaders under '{yellow_header}':")
#                 for sh in subheaders:
#                     print(f"  - {sh}")

# if __name__ == "__main__":
#     main()




import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

file_path = 'rebar_dom.xlsx'  # Replace with your file path

YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

def is_yellow(cell):
    fill = cell.fill
    if fill.fill_type == 'solid':
        color = fill.start_color
        if color.rgb:
            return color.rgb.upper() in YELLOW_RGB_VALUES
        if color.indexed is not None:
            return color.indexed == 6
    return False

def find_yellow_header_row(ws, max_rows=10):
    for row_num in range(1, max_rows+1):
        row = ws[row_num]
        if any(is_yellow(cell) for cell in row):
            return row_num
    return None

def get_merged_range_for_cell(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

def main():
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing Sheet: {sheet_name}")
        
        header_row_num = find_yellow_header_row(ws, max_rows=10)
        if not header_row_num:
            print(f"No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
            continue
        
        subheader_row_num = header_row_num + 1
        
        # Load whole sheet into pandas first (so we can slice columns easily)
        # Note: pandas reads by default with header=0, but here header is at subheader_row_num - 1 (zero-based)
        # Adjust header param accordingly
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=subheader_row_num-1)
        
        # This df columns are the subheaders row in the Excel sheet
        
        header_row = ws[header_row_num]
        
        # Prepare Excel writer to save all yellow header groups as separate sheets inside one Excel file
        output_filename = f"{sheet_name}_extracted.xlsx"
        writer = pd.ExcelWriter(output_filename, engine='openpyxl')
        
        for cell in header_row:
            if is_yellow(cell):
                merged_range = get_merged_range_for_cell(ws, cell)
                if merged_range:
                    min_col = merged_range.min_col
                    max_col = merged_range.max_col
                else:
                    min_col = cell.col_idx
                    max_col = cell.col_idx
                
                yellow_header = cell.value
                print(f"\nExtracting columns under yellow header: '{yellow_header}' spanning cols {min_col} to {max_col}")
                
                # Get subheaders for these columns to use as pandas columns to extract
                subheaders = []
                for col_idx in range(min_col, max_col + 1):
                    sub_cell = ws.cell(row=subheader_row_num, column=col_idx)
                    subheaders.append(sub_cell.value)
                
                # Filter dataframe columns matching these subheaders (some might be None or NaN)
                subheaders_filtered = [sh for sh in subheaders if sh in df.columns and sh is not None]
                
                if not subheaders_filtered:
                    print(f"No matching subheader columns found in dataframe for header '{yellow_header}'. Skipping.")
                    continue
                
                extracted_df = df[subheaders_filtered]
                
                # Write extracted columns to a sheet named by the yellow header (cleaned for filename)
                safe_sheet_name = str(yellow_header)[:30]  # Excel sheet name limit
                safe_sheet_name = safe_sheet_name.replace('/', '_').replace('\\', '_')
                
                extracted_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                print(f"Saved {len(subheaders_filtered)} columns under '{yellow_header}' to sheet '{safe_sheet_name}'")
        
        writer.close()
        print(f"All extracted sheets saved to {output_filename}")

if __name__ == "__main__":
    main()
