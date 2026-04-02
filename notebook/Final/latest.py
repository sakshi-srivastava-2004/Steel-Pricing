# import openpyxl
# import pandas as pd
# import re

# file_path = 'pig iron_D_dom.xlsx'

# YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

# def is_yellow(cell):
#     fill = cell.fill
#     if fill.fill_type == 'solid':
#         color = fill.start_color
#         if color.rgb:
#             return color.rgb.upper() in YELLOW_RGB_VALUES
#         if color.indexed is not None:
#             return color.indexed == 6
#     return False

# def find_yellow_header_row(ws, max_rows=10):
#     for row_num in range(1, max_rows + 1):
#         row = ws[row_num]
#         if any(is_yellow(cell) for cell in row):
#             return row_num
#     return None

# def get_merged_range_for_cell(ws, cell):
#     for merged_range in ws.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             return merged_range
#     return None

# def sanitize_sheet_name(name):
#     name = str(name).strip()
#     name = name.replace('&', 'and')
#     name = re.sub(r'[:\\/*?\[\]]', '_', name)
#     return name[:31]

# def calculate_weekly_averages(df):
#     weekly_averages = []
#     h_indices = df[df.iloc[:, 0] == 'H'].index.to_list()
#     h_indices.append(len(df))  # Add last index to close final block
#     start_idx = 0
#     week_num = 1

#     for h_idx in h_indices:
#         if start_idx < h_idx:
#             block = df.iloc[start_idx:h_idx]
#             avg_row = {}
#             for col in df.columns:
#                 values = pd.to_numeric(block[col], errors='coerce').dropna()
#                 avg_row[col] = values.mean() if not values.empty else None
#             avg_row["Week"] = week_num
#             weekly_averages.append(avg_row)
#             week_num += 1
#         start_idx = h_idx + 1

#     weekly_avg_df = pd.DataFrame(weekly_averages)
#     cols = ["Week"] + [col for col in df.columns if col != "Week"]
#     return weekly_avg_df[cols]

# def main():
#     wb = openpyxl.load_workbook(file_path, data_only=True)

#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         print(f"\nProcessing Sheet: {sheet_name}")

#         header_row_num = find_yellow_header_row(ws)
#         if not header_row_num:
#             print(f"No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
#             continue

#         subheader_row_num = header_row_num + 1
#         max_col = ws.max_column
#         max_row = ws.max_row

#         output_filename = f"{sanitize_sheet_name(sheet_name)}_weekly_avg.xlsx"
#         writer = pd.ExcelWriter(output_filename, engine='xlsxwriter')
#         start_row = 0

#         header_row = ws[header_row_num]
#         processed_headers = set()

#         for cell in header_row:
#             if is_yellow(cell) and cell.value and cell.value not in processed_headers:
#                 material = str(cell.value).strip()
#                 processed_headers.add(material)

#                 merged_range = get_merged_range_for_cell(ws, cell)
#                 min_col = merged_range.min_col if merged_range else cell.col_idx
#                 max_col = merged_range.max_col if merged_range else cell.col_idx

#                 subheaders = []
#                 for col in range(min_col, max_col + 1):
#                     sub_val = ws.cell(row=subheader_row_num, column=col).value
#                     if sub_val:
#                         subheaders.append(str(sub_val).strip())

#                 if not subheaders:
#                     print(f"No subheaders found for '{material}'. Skipping.")
#                     continue

#                 data = {subheader: [] for subheader in subheaders}
#                 row = subheader_row_num + 1
#                 while row <= max_row:
#                     is_empty_row = True
#                     for i, col in enumerate(range(min_col, max_col + 1)):
#                         val = ws.cell(row=row, column=col).value
#                         if val is not None:
#                             is_empty_row = False
#                         data[subheaders[i]].append(val)
#                     if is_empty_row:
#                         break
#                     row += 1

#                 df = pd.DataFrame(data).dropna(how='all')
#                 if df.empty:
#                     continue

#                 # Calculate weekly averages
#                 weekly_avg_df = calculate_weekly_averages(df)

#                 # Add header/title
#                 title = f"{material} ({file_path})"
#                 df_with_title = pd.DataFrame({material: [title]})
#                 df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)

#                 # Write weekly averages
#                 weekly_avg_df.to_excel(writer, startrow=start_row + 1, index=False)
#                 print(f"→ Weekly averages for '{material}' written with {len(weekly_avg_df)} weeks")
#                 start_row += len(weekly_avg_df) + 3  # Leave space

#         writer.close()
#         print(f"✅ Saved to file: {output_filename}")

# if __name__ == "__main__":
#     main()



# import openpyxl
# import pandas as pd
# import re

# file_path = 'hrc_W_dom.xlsx'

# YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

# def is_yellow(cell):
#     fill = cell.fill
#     if fill.fill_type == 'solid':
#         color = fill.start_color
#         if color.rgb:
#             return color.rgb.upper() in YELLOW_RGB_VALUES
#         if color.indexed is not None:
#             return color.indexed == 6
#     return False

# def find_yellow_header_row(ws, max_rows=10):
#     for row_num in range(1, max_rows + 1):
#         row = ws[row_num]
#         if any(is_yellow(cell) for cell in row):
#             return row_num
#     return None

# def get_merged_range_for_cell(ws, cell):
#     for merged_range in ws.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             return merged_range
#     return None

# def sanitize_sheet_name(name):
#     name = str(name).strip()
#     name = name.replace('&', 'and')
#     name = re.sub(r'[:\\/*?\[\]]', '_', name)
#     return name[:31]

# def calculate_weekly_averages(df):
#     weekly_averages = []
#     h_indices = df[df.iloc[:, 0] == 'H'].index.to_list()
#     h_indices.append(len(df))
#     start_idx = 0
#     week_num = 1

#     for h_idx in h_indices:
#         if start_idx < h_idx:
#             block = df.iloc[start_idx:h_idx]
#             avg_row = {}
#             for col in df.columns:
#                 values = pd.to_numeric(block[col], errors='coerce').dropna()
#                 avg_row[col] = values.mean() if not values.empty else None
#             avg_row["Week"] = week_num
#             weekly_averages.append(avg_row)
#             week_num += 1
#         start_idx = h_idx + 1

#     weekly_avg_df = pd.DataFrame(weekly_averages)
#     cols = ["Week"] + [col for col in df.columns if col != "Week"]
#     return weekly_avg_df[cols]

# def main():
#     wb = openpyxl.load_workbook(file_path, data_only=True)

#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         print(f"\nProcessing Sheet: {sheet_name}")

#         header_row_num = find_yellow_header_row(ws)
#         if not header_row_num:
#             print(f"No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
#             continue

#         subheader_row_num = header_row_num + 1
#         max_col = ws.max_column
#         max_row = ws.max_row

#         output_filename = f"{sanitize_sheet_name(sheet_name)}_weekly_avg.xlsx"
#         writer = pd.ExcelWriter(output_filename, engine='xlsxwriter')
#         start_row = 0

#         header_row = ws[header_row_num]
#         processed_headers = set()

#         for cell in header_row:
#             if is_yellow(cell) and cell.value and cell.value not in processed_headers:
#                 material = str(cell.value).strip()
#                 processed_headers.add(material)

#                 merged_range = get_merged_range_for_cell(ws, cell)
#                 min_col = merged_range.min_col if merged_range else cell.col_idx
#                 max_col = merged_range.max_col if merged_range else cell.col_idx

#                 subheaders = []
#                 for col in range(min_col, max_col + 1):
#                     sub_val = ws.cell(row=subheader_row_num, column=col).value
#                     if sub_val:
#                         subheaders.append(str(sub_val).strip())

#                 if not subheaders:
#                     print(f"No subheaders found for '{material}'. Skipping.")
#                     continue

#                 data = {subheader: [] for subheader in subheaders}
#                 row = subheader_row_num + 1
#                 while row <= max_row:
#                     is_empty_row = True
#                     for i, col in enumerate(range(min_col, max_col + 1)):
#                         val = ws.cell(row=row, column=col).value
#                         if val is not None:
#                             is_empty_row = False
#                         data[subheaders[i]].append(val)
#                     if is_empty_row:
#                         break
#                     row += 1

#                 df = pd.DataFrame(data).dropna(how='all')
#                 if df.empty:
#                     continue

#                 # Normalize column names
#                 normalized_cols = [str(c).strip().lower() for c in df.columns]
#                 has_week = any(col == "week" for col in normalized_cols)
#                 has_date = any("date" in col for col in normalized_cols)

#                 if has_week and not has_date:
#                     # Already weekly data – save as-is
#                     df_with_title = pd.DataFrame({material: [f"{material} (Pre-averaged — Week column)"]})
#                     df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)
#                     df.to_excel(writer, startrow=start_row + 1, index=False)
#                     print(f"→ Detected 'Week' only — skipped averaging for '{material}'")
#                     start_row += len(df) + 3
#                     continue

#                 elif has_date:
#                     # Average using date column
#                     weekly_avg_df = calculate_weekly_averages(df)
#                     df_with_title = pd.DataFrame({material: [f"{material} (Calculated from Date)"]})
#                     df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)
#                     weekly_avg_df.to_excel(writer, startrow=start_row + 1, index=False)
#                     print(f"→ Detected 'Date' — calculated averages for '{material}'")
#                     start_row += len(weekly_avg_df) + 3
#                     continue

#                 else:
#                     # Average based on 'H' separation
#                     weekly_avg_df = calculate_weekly_averages(df)
#                     df_with_title = pd.DataFrame({material: [f"{material} (Calculated from H)"]})
#                     df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)
#                     weekly_avg_df.to_excel(writer, startrow=start_row + 1, index=False)
#                     print(f"→ No 'Date' or 'Week' — calculated averages for '{material}' using H")
#                     start_row += len(weekly_avg_df) + 3

#         writer.close()
#         print(f"✅ Saved to file: {output_filename}")

# if __name__ == "__main__":
#     main()



#  week not working

# import openpyxl
# import pandas as pd
# import re

# file_path = 'hrc_W_dom.xlsx'

# YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

# def is_yellow(cell):
#     fill = cell.fill
#     if fill.fill_type == 'solid':
#         color = fill.start_color
#         if color.rgb:
#             return color.rgb.upper() in YELLOW_RGB_VALUES
#         if color.indexed is not None:
#             return color.indexed == 6
#     return False

# def find_yellow_header_row(ws, max_rows=10):
#     for row_num in range(1, max_rows + 1):
#         row = ws[row_num]
#         if any(is_yellow(cell) for cell in row):
#             return row_num
#     return None

# def get_merged_range_for_cell(ws, cell):
#     for merged_range in ws.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             return merged_range
#     return None

# def sanitize_sheet_name(name):
#     name = str(name).strip()
#     name = name.replace('&', 'and')
#     name = re.sub(r'[:\\/*?\[\]]', '_', name)
#     return name[:31]

# def calculate_weekly_averages(df):
#     weekly_averages = []
#     # Check if the first column contains 'H'
#     if df.iloc[:, 0].astype(str).str.upper().str.strip().eq('H').any():
#         h_indices = df[df.iloc[:, 0].astype(str).str.upper().str.strip() == 'H'].index.to_list()
#         h_indices.append(len(df))
#         start_idx = 0
#         week_num = 1

#         for h_idx in h_indices:
#             if start_idx < h_idx:
#                 block = df.iloc[start_idx:h_idx]
#                 avg_row = {}
#                 for col in df.columns:
#                     values = pd.to_numeric(block[col], errors='coerce').dropna()
#                     avg_row[col] = values.mean() if not values.empty else None
#                 avg_row["Week"] = week_num
#                 weekly_averages.append(avg_row)
#                 week_num += 1
#             start_idx = h_idx + 1
#     else: # If 'H' is not present, assume each row is a week or block
#         avg_row = {}
#         for col in df.columns:
#             values = pd.to_numeric(df[col], errors='coerce').dropna()
#             avg_row[col] = values.mean() if not values.empty else None
#         avg_row["Week"] = 1 # Assign week 1 if no 'H'
#         weekly_averages.append(avg_row)

#     weekly_avg_df = pd.DataFrame(weekly_averages)
#     cols = ["Week"] + [col for col in df.columns if col != "Week"]
#     # Ensure 'Week' is the first column if it exists
#     if "Week" in weekly_avg_df.columns:
#         return weekly_avg_df[cols]
#     return weekly_avg_df


# def main():
#     wb = openpyxl.load_workbook(file_path, data_only=True)

#     for sheet_name in wb.sheetnames:
#         ws = wb[sheet_name]
#         print(f"\nProcessing Sheet: {sheet_name}")

#         header_row_num = find_yellow_header_row(ws)
#         if not header_row_num:
#             print(f"No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
#             continue

#         subheader_row_num = header_row_num + 1
#         max_row = ws.max_row

#         output_filename = f"{sanitize_sheet_name(sheet_name)}_weekly_avg.xlsx"
#         writer = pd.ExcelWriter(output_filename, engine='xlsxwriter')
#         start_row = 0

#         header_row = ws[header_row_num]
#         processed_headers = set()

#         for cell in header_row:
#             if is_yellow(cell) and cell.value and cell.value not in processed_headers:
#                 material = str(cell.value).strip()
#                 processed_headers.add(material)

#                 merged_range = get_merged_range_for_cell(ws, cell)
#                 min_col = merged_range.min_col if merged_range else cell.col_idx
#                 max_col = merged_range.max_col if merged_range else cell.col_idx

#                 subheaders = []
#                 for col in range(min_col, max_col + 1):
#                     sub_val = ws.cell(row=subheader_row_num, column=col).value
#                     if sub_val:
#                         subheaders.append(str(sub_val).strip())

#                 if not subheaders:
#                     print(f"No subheaders found for '{material}'. Skipping.")
#                     continue

#                 data = {subheader: [] for subheader in subheaders}
#                 row = subheader_row_num + 1
#                 while row <= max_row:
#                     is_empty_row = True
#                     # Collect row data, checking if the entire row is empty
#                     row_data = [ws.cell(row=row, column=col).value for col in range(min_col, max_col + 1)]
#                     if any(val is not None for val in row_data):
#                         is_empty_row = False
#                         for i, val in enumerate(row_data):
#                             data[subheaders[i]].append(val)
#                     if is_empty_row:
#                         break # Stop if an entire row is empty
#                     row += 1

#                 df = pd.DataFrame(data).dropna(how='all')
#                 if df.empty:
#                     continue

#                 # Normalize column names for checking
#                 normalized_cols = [str(c).strip().lower() for c in df.columns]
                
#                 # Check for "week" in any of the column names
#                 has_week_column = any("week" in col for col in normalized_cols)
#                 has_date_column = any("date" in col for col in normalized_cols)

#                 if has_week_column:
#                     # If 'Week' column exists, save the data as is without averaging
#                     df_with_title = pd.DataFrame({material: [f"{material} (Directly Extracted - Week column present)"]})
#                     df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)
#                     df.to_excel(writer, startrow=start_row + 1, index=False)
#                     print(f"→ Detected 'Week' column — directly extracted data for '{material}'")
#                     start_row += len(df) + 3 # Add 3 for title, data, and a blank row
#                     continue
#                 elif has_date_column:
#                     # If 'Date' column exists, calculate weekly averages
#                     weekly_avg_df = calculate_weekly_averages(df)
#                     df_with_title = pd.DataFrame({material: [f"{material} (Calculated from Date)"]})
#                     df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)
#                     weekly_avg_df.to_excel(writer, startrow=start_row + 1, index=False)
#                     print(f"→ Detected 'Date' column — calculated averages for '{material}'")
#                     start_row += len(weekly_avg_df) + 3
#                 else:
#                     # Otherwise, calculate weekly averages based on 'H' separation
#                     weekly_avg_df = calculate_weekly_averages(df)
#                     df_with_title = pd.DataFrame({material: [f"{material} (Calculated from H)"]})
#                     df_with_title.to_excel(writer, startrow=start_row, index=False, header=False)
#                     weekly_avg_df.to_excel(writer, startrow=start_row + 1, index=False)
#                     print(f"→ No 'Date' or 'Week' column — calculated averages for '{material}' using H")
#                     start_row += len(weekly_avg_df) + 3

#         writer.close()
#         print(f"✅ Saved to file: {output_filename}")

# if __name__ == "__main__":
#     main()



import openpyxl
import pandas as pd
import re

file_path = 'ingot_D_dom.xlsx'

YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'}

def is_yellow(cell):
    fill = cell.fill
    if fill.fill_type == 'solid':
        color = fill.start_color
        if color.rgb:
            return color.rgb.upper() in YELLOW_RGB_VALUES
        if color.indexed is not None:
            return color.indexed == 6
    return False

def find_yellow_header_row(ws, max_rows=10):
    for row_num in range(1, max_rows + 1):
        row = ws[row_num]
        if any(is_yellow(cell) for cell in row):
            return row_num
    return None

def get_merged_range_for_cell(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

def sanitize_sheet_name(name):
    name = str(name).strip()
    name = name.replace('&', 'and')
    name = re.sub(r'[:\\/*?\[\]]', '_', name)
    return name[:31]

def calculate_weekly_averages(df):
    # Normalize column names for internal checks within this function
    normalized_cols_df = [str(col).strip().lower() for col in df.columns]

    # If a 'week' column already exists (case-insensitive), return the DataFrame as is.
    # This scenario should ideally be handled before calling this function,
    # but this is a safeguard.
    if any("Week" in col for col in normalized_cols_df):
        print("    (Note: 'Week' column already present, skipping averaging in calculate_weekly_averages)")
        return df

    weekly_averages = []
    
    # Check if the first column contains 'H' (case-insensitive and trimmed)
    # Ensure to handle non-string types gracefully by converting to string first
    first_col_values = df.iloc[:, 0].astype(str).str.strip().str.upper()

    if first_col_values.eq('H').any():
        h_indices = first_col_values[first_col_values == 'H'].index.to_list()
        h_indices.append(len(df)) # Add the end of the DataFrame as a block terminator
        
        start_idx = 0
        week_num = 1

        for h_idx in h_indices:
            if start_idx < h_idx:
                block = df.iloc[start_idx:h_idx]
                avg_row = {}
                for col in df.columns:
                    # Convert to numeric, coerce errors to NaN, then drop NaNs for mean calculation
                    values = pd.to_numeric(block[col], errors='coerce').dropna()
                    avg_row[col] = values.mean() if not values.empty else None
                avg_row["Week"] = week_num
                weekly_averages.append(avg_row)
                week_num += 1
            start_idx = h_idx + 1 # Move to the row after 'H' for the next block
    else:
        # If no 'H' marker is found, consider the entire DataFrame as one block for averaging,
        # or if there's a 'Date' column, handle that.
        # This part needs to align with how 'Date' is handled in main.
        # For simplicity here, if no 'H', just average the whole thing as Week 1
        # if not explicitly handled by 'Date' in the main loop.
        avg_row = {}
        for col in df.columns:
            values = pd.to_numeric(df[col], errors='coerce').dropna()
            avg_row[col] = values.mean() if not values.empty else None
        avg_row["Week"] = 1 # Assign week 1 if no 'H' marker
        weekly_averages.append(avg_row)

    weekly_avg_df = pd.DataFrame(weekly_averages)
    
    # Ensure 'Week' is the first column if it was just created
    if "Week" in weekly_avg_df.columns:
        cols_order = ["Week"] + [col for col in df.columns if col != "Week"]
        return weekly_avg_df[cols_order]
    return weekly_avg_df


def main():
    wb = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing Sheet: {sheet_name}")

        header_row_num = find_yellow_header_row(ws)
        if not header_row_num:
            print(f"No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
            continue

        subheader_row_num = header_row_num + 1
        max_row = ws.max_row

        output_filename = f"{sanitize_sheet_name(sheet_name)}_processed.xlsx" # Changed filename for clarity
        writer = pd.ExcelWriter(output_filename, engine='xlsxwriter')
        start_row = 0

        header_row_cells = ws[header_row_num] # Get all cells in the header row
        processed_headers = set() # To keep track of already processed yellow headers

        # Iterate through columns to find yellow headers and their corresponding subheaders/data
        for col_idx in range(1, ws.max_column + 1): # Iterate through all possible columns
            cell = ws.cell(row=header_row_num, column=col_idx)

            if is_yellow(cell) and cell.value and str(cell.value).strip() not in processed_headers:
                material = str(cell.value).strip()
                processed_headers.add(material)

                merged_range = get_merged_range_for_cell(ws, cell)
                min_data_col = merged_range.min_col if merged_range else cell.col_idx
                max_data_col = merged_range.max_col if merged_range else cell.col_idx

                subheaders = []
                # Collect subheaders for the current material block
                for sub_col_idx in range(min_data_col, max_data_col + 1):
                    sub_val = ws.cell(row=subheader_row_num, column=sub_col_idx).value
                    if sub_val is not None: # Ensure subheader is not None
                        subheaders.append(str(sub_val).strip())
                    else:
                        # If a subheader is empty, assign a placeholder or skip
                        subheaders.append(f"Unnamed_Col_{sub_col_idx}")


                if not subheaders or all(s.startswith("Unnamed_Col_") for s in subheaders):
                    print(f"No meaningful subheaders found for '{material}'. Skipping.")
                    continue

                data_rows = []
                row_iterator = subheader_row_num + 1
                while row_iterator <= max_row:
                    current_row_data = []
                    is_empty_row = True
                    # Collect data for the current material block
                    for data_col_idx in range(min_data_col, max_data_col + 1):
                        val = ws.cell(row=row_iterator, column=data_col_idx).value
                        current_row_data.append(val)
                        if val is not None and str(val).strip() != '': # Check for non-empty values
                            is_empty_row = False
                    
                    if is_empty_row:
                        break # Stop when an entire row in the data block is empty
                    data_rows.append(current_row_data)
                    row_iterator += 1

                # Create DataFrame from collected data and subheaders
                # Ensure subheaders list matches the number of columns in data_rows
                if data_rows:
                    df = pd.DataFrame(data_rows, columns=subheaders)
                else:
                    df = pd.DataFrame(columns=subheaders) # Create empty DF with headers if no data

                # Drop rows where all values are NaN after conversion, or all are empty strings
                df = df.dropna(how='all')
                if df.empty:
                    print(f"No valid data found under '{material}'. Skipping.")
                    continue

                # Normalize column names for detection
                normalized_df_cols = [str(col).strip().lower() for col in df.columns]
                
                has_week_column = any("Week" in col for col in normalized_df_cols)
                has_date_column = any("Date" in col for col in normalized_df_cols)

                if has_week_column:
                    # If 'Week' column exists, save the data as is without averaging
                    df_with_title = pd.DataFrame({material: [f"{material} (Directly Extracted - Week column present)"]})
                    df_with_title.to_excel(writer, sheet_name='Output', startrow=start_row, index=False, header=False)
                    df.to_excel(writer, sheet_name='Output', startrow=start_row + 1, index=False)
                    print(f"→ Detected 'Week' column — directly extracted data for '{material}'")
                    start_row += len(df) + 3 # Add 3 for title, data, and a blank row
                elif has_date_column:
                    # If 'Date' column exists, calculate weekly averages
                    weekly_avg_df = calculate_weekly_averages(df)
                    df_with_title = pd.DataFrame({material: [f"{material} (Calculated from Date)"]})
                    df_with_title.to_excel(writer, sheet_name='Output', startrow=start_row, index=False, header=False)
                    weekly_avg_df.to_excel(writer, sheet_name='Output', startrow=start_row + 1, index=False)
                    print(f"→ Detected 'Date' column — calculated averages for '{material}'")
                    start_row += len(weekly_avg_df) + 3
                else:
                    # Otherwise, calculate weekly averages based on 'H' separation
                    weekly_avg_df = calculate_weekly_averages(df)
                    df_with_title = pd.DataFrame({material: [f"{material} (Calculated from H)"]})
                    df_with_title.to_excel(writer, sheet_name='Output', startrow=start_row, index=False, header=False)
                    weekly_avg_df.to_excel(writer, sheet_name='Output', startrow=start_row + 1, index=False)
                    print(f"→ No 'Date' or 'Week' column — calculated averages for '{material}' using H")
                    start_row += len(weekly_avg_df) + 3

        writer.close()
        print(f"✅ Saved to file: {output_filename}")

if __name__ == "__main__":
    main()