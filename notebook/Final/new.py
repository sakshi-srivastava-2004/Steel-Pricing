import openpyxl
import pandas as pd
import re
import numpy as np
import os
import glob

# --- Configuration ---
# IMPORTANT: Change this to the actual directory path where your Excel files are located
input_directory = 'C:/intership/notebook/input' 
# The name of the single consolidated output Excel file
consolidated_output_file_name = os.path.join(input_directory, 'consolidated_processed_data.xlsx')

# RGB values for yellow color used in Excel for header detection
YELLOW_RGB_VALUES = {'FFFFFF00', 'FFFF00'} # Add other common yellow RGBs if needed

# --- Utility Functions ---

def is_yellow(cell):
    """Checks if a cell's fill color is yellow."""
    fill = cell.fill
    if fill.fill_type == 'solid':
        color = fill.start_color
        if color.rgb:
            return color.rgb.upper() in YELLOW_RGB_VALUES
        if color.indexed is not None:
            # Indexed color 6 is often yellow in default palettes
            return color.indexed == 6
    return False

def find_yellow_header_row(ws, max_rows=10):
    """Finds the row number of the first row containing any yellow-highlighted cell."""
    for row_num in range(1, max_rows + 1):
        row = ws[row_num]
        if any(is_yellow(cell) for cell in row):
            return row_num
    return None

def get_merged_range_for_cell(ws, cell):
    """
    Returns the merged cell range object if the given cell is part of one,
    otherwise returns None.
    """
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

def sanitize_sheet_name(name):
    """
    Sanitizes a string to be a valid Excel sheet name (max 31 chars, no invalid chars).
    """
    name = str(name).strip()
    # Replace invalid characters with underscore
    name = name.replace('&', 'and') # A common replacement
    name = re.sub(r'[:\\/*?\[\]]', '_', name)
    # Truncate to 31 characters, which is Excel's limit
    return name[:31]

def average_weeks_only_logic(df_block):
    """
    Applies H-detection and averaging logic to a DataFrame block.
    Dynamically finds the column containing 'H' markers and averages data before each 'H'.
    """
    # Replace common empty/placeholder strings with NaN for proper numeric conversion
    df_block = df_block.replace(['-', '', ' '], np.nan).reset_index(drop=True)

    averaged_weeks = []
    current_block_start_idx = 0
    week_number = 1

    h_marker_col_name = None
    
    # Identify the column containing 'H' markers
    normalized_cols_check = [str(col).strip().lower() for col in df_block.columns]
    
    for col_idx, col_name in enumerate(df_block.columns):
        # Exclude 'date' and 'week' columns from the 'H' search as 'H' is a data delimiter
        if "date" not in normalized_cols_check[col_idx] and "week" not in normalized_cols_check[col_idx]:
            col_series = df_block[col_name].astype(str).str.strip().str.upper()
            if col_series.eq('H').any(): # Check if 'H' exists in this column
                h_marker_col_name = col_name
                break # Found the column with 'H' markers, stop searching
    
    # If no 'H' marker is found in any eligible data column, return an empty DataFrame
    # This signals that H-based averaging cannot be performed.
    if h_marker_col_name is None:
        return pd.DataFrame(columns=df_block.columns) 

    h_marker_series = df_block[h_marker_col_name].astype(str).str.strip().str.upper()
    
    # Get the row indices where 'H' markers are present. These define block boundaries.
    h_indices = h_marker_series[h_marker_series == 'H'].index.to_list()
    
    # Process blocks delimited by 'H' markers
    for h_idx in h_indices:
        # A block consists of data from `current_block_start_idx` up to (but not including) `h_idx`
        block = df_block.iloc[current_block_start_idx:h_idx]
        
        if not block.empty:
            avg_row = {}
            for col in df_block.columns:
                if "week" in str(col).strip().lower() or "date" in str(col).strip().lower():
                    # For 'Week' or 'Date' columns, take the first non-null value in the block
                    first_val = block[col].dropna().iloc[0] if not block[col].dropna().empty else None
                    avg_row[col] = first_val
                else:
                    # For other columns, attempt numeric conversion and calculate the mean
                    values = pd.to_numeric(block[col], errors='coerce').dropna()
                    avg_row[col] = values.mean() if not values.empty else np.nan 
            
            avg_row["Calculated_Week"] = week_number 
            averaged_weeks.append(avg_row)
            week_number += 1
        
        # The next block starts immediately after the 'H' marker
        current_block_start_idx = h_idx + 1

    # Handle any remaining data after the last 'H' marker
    # This also covers the case where no 'H' markers were found at all in the DataFrame.
    remaining_block = df_block.iloc[current_block_start_idx:]
    if not remaining_block.empty:
        avg_row = {}
        for col in df_block.columns:
            if "week" in str(col).strip().lower() or "date" in str(col).strip().lower():
                first_val = remaining_block[col].dropna().iloc[0] if not remaining_block[col].dropna().empty else None
                avg_row[col] = first_val
            else:
                values = pd.to_numeric(remaining_block[col], errors='coerce').dropna()
                avg_row[col] = values.mean() if not values.empty else np.nan
        
        avg_row["Calculated_Week"] = week_number 
        averaged_weeks.append(avg_row)
        
    result_df = pd.DataFrame(averaged_weeks)
    
    # Reorder columns to place 'Calculated_Week' at the beginning for consistent output
    if 'Calculated_Week' in result_df.columns:
        cols = ['Calculated_Week'] + [col for col in result_df.columns if col != 'Calculated_Week']
        result_df = result_df[cols]

    return result_df

def calculate_weekly_averages(df):
    """
    Main dispatch function for averaging logic.
    Checks for an existing 'Week' column first. If not found, attempts H-based averaging.
    """
    normalized_cols_df = [str(col).strip().lower() for col in df.columns]

    # If a 'week' column already exists (case-insensitive), return the DataFrame as is.
    # No averaging needed if weekly data is already provided.
    if any("week" in col for col in normalized_cols_df):
        return df

    # Create a copy to avoid modifying the original DataFrame passed in
    df_cleaned = df.replace(['-', '', ' '], np.nan).copy()

    # Call the H-detection and averaging logic
    averaged_df = average_weeks_only_logic(df_cleaned)

    # If average_weeks_only_logic returned an empty DataFrame, it means no 'H' was found for grouping.
    # In this case, return the original DataFrame if it contained data.
    if averaged_df.empty and df_cleaned.shape[0] > 0: 
        return df 
    
    return averaged_df

# --- Main Processing Logic ---

def main():
    # Get a list of all .xlsx files in the specified input directory
    excel_files = glob.glob(os.path.join(input_directory, '*.xlsx'))

    if not excel_files:
        print(f"No Excel files found in the directory: {input_directory}")
        return

    print(f"Found {len(excel_files)} Excel files to process in {input_directory}:")
    for file_path in excel_files:
        print(f"- {os.path.basename(file_path)}")

    # This dictionary will store all processed data, grouped by their original sheet name.
    # Each value will be a list of (title_DataFrame, data_DataFrame) pairs.
    consolidated_output_data = {} 

    for file_path in excel_files:
        print(f"\n--- Processing file: {os.path.basename(file_path)} ---")
        try:
            # Load the workbook in data_only mode to get cell values, not formulas
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            print(f"Error: The file '{os.path.basename(file_path)}' was not found. Skipping.")
            continue
        except Exception as e:
            print(f"An error occurred while opening '{os.path.basename(file_path)}': {e}. Skipping.")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Processing Sheet: {sheet_name}")

            header_row_num = find_yellow_header_row(ws)
            if not header_row_num:
                print(f"  No yellow highlighted header row found in sheet '{sheet_name}'. Skipping.")
                continue

            subheader_row_num = header_row_num + 1
            max_row = ws.max_row

            # Keep track of processed material headers to avoid duplicate processing in merged cells
            processed_headers = set()

            min_overall_col = 1
            max_overall_col = ws.max_column
            
            # Identify the general 'Week' or 'Date' column position, if they exist in the headers
            week_col_idx = None
            date_col_idx = None
            
            for col_idx_check in range(min_overall_col, max_overall_col + 1):
                cell_val_header = str(ws.cell(row=header_row_num, column=col_idx_check).value).strip().lower()
                cell_val_subheader = str(ws.cell(row=subheader_row_num, column=col_idx_check).value).strip().lower()

                if "week" in cell_val_header and week_col_idx is None:
                    week_col_idx = col_idx_check
                if "week" in cell_val_subheader and week_col_idx is None:
                    week_col_idx = col_idx_check

                if "date" in cell_val_header and date_col_idx is None:
                    date_col_idx = col_idx_check
                if "date" in cell_val_subheader and date_col_idx is None:
                    date_col_idx = col_idx_check
            
            has_week_column_overall = week_col_idx is not None
            has_date_column_overall = date_col_idx is not None

            # Iterate through columns to find material blocks (yellow headers)
            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row_num, column=col_idx)

                # Process if it's a yellow header and not already processed (due to merged cells)
                if is_yellow(header_cell) and header_cell.value and str(header_cell.value).strip() not in processed_headers:
                    material = str(header_cell.value).strip()
                    processed_headers.add(material)

                    # Determine the column range for the current material block
                    merged_range = get_merged_range_for_cell(ws, header_cell)
                    min_data_col_block = merged_range.min_col if merged_range else header_cell.col_idx
                    max_data_col_block = merged_range.max_col if merged_range else header_cell.col_idx

                    cols_to_extract_indices = []
                    final_column_names = []

                    # Add Week/Date columns if they exist and are not already part of the material block
                    if has_week_column_overall and week_col_idx not in cols_to_extract_indices:
                        cols_to_extract_indices.append(week_col_idx)
                        col_name = str(ws.cell(row=header_row_num, column=week_col_idx).value).strip()
                        if not col_name: # Fallback to subheader if main header is empty
                            col_name = str(ws.cell(row=subheader_row_num, column=week_col_idx).value).strip()
                        final_column_names.append(col_name or f"Week_Col_{week_col_idx}")
                    
                    if has_date_column_overall and date_col_idx != week_col_idx and date_col_idx not in cols_to_extract_indices:
                        cols_to_extract_indices.append(date_col_idx)
                        col_name = str(ws.cell(row=header_row_num, column=date_col_idx).value).strip()
                        if not col_name:
                            col_name = str(ws.cell(row=subheader_row_num, column=date_col_idx).value).strip()
                        final_column_names.append(col_name or f"Date_Col_{date_col_idx}")

                    # Add columns specific to the current material block
                    for current_col_idx in range(min_data_col_block, max_data_col_block + 1):
                        if current_col_idx not in cols_to_extract_indices: # Avoid duplicating Week/Date if they fall into the material range
                            sub_val = ws.cell(row=subheader_row_num, column=current_col_idx).value
                            if sub_val is not None and str(sub_val).strip() != '':
                                final_column_names.append(str(sub_val).strip())
                                cols_to_extract_indices.append(current_col_idx)
                            else:
                                final_column_names.append(f"Unnamed_Col_{current_col_idx}")
                                cols_to_extract_indices.append(current_col_idx)

                    # Ensure columns are in the correct order (matching their original column index)
                    combined_cols_info = sorted(zip(cols_to_extract_indices, final_column_names))
                    actual_df_columns_order = [name for idx, name in combined_cols_info]
                    actual_cols_indices_order = [idx for idx, name in combined_cols_info]

                    if not actual_df_columns_order or all(s.startswith("Unnamed_Col_") for s in actual_df_columns_order):
                        print(f"    No meaningful columns identified for '{material}'. Skipping.")
                        continue

                    # Extract data rows for the current material block
                    data_rows = []
                    row_iterator = subheader_row_num + 1
                    while row_iterator <= max_row:
                        current_row_data = []
                        is_empty_row = True
                        for data_col_idx in actual_cols_indices_order:
                            val = ws.cell(row=row_iterator, column=data_col_idx).value
                            current_row_data.append(val)
                            if val is not None and str(val).strip() != '':
                                is_empty_row = False # Row is not empty if any cell has content
                        
                        if is_empty_row: # Stop if an entirely empty row is encountered
                            break
                        data_rows.append(current_row_data)
                        row_iterator += 1

                    if data_rows:
                        df = pd.DataFrame(data_rows, columns=actual_df_columns_order)
                    else:
                        df = pd.DataFrame(columns=actual_df_columns_order)

                    # Drop rows where all values are NaN (fully empty rows after conversion)
                    df = df.dropna(how='all')
                    if df.empty:
                        print(f"    No valid data found under '{material}'. Skipping.")
                        continue

                    normalized_df_cols = [str(col).strip().lower() for col in df.columns]
                    
                    # Determine how to process and title the data block
                    processed_data_to_store = None
                    material_title = ""

                    if has_week_column_overall:
                        # If a 'Week' column was found in the overall headers, just extract directly
                        material_title = f"{material} (From file: {os.path.basename(file_path)} - Directly Extracted - Week column present)"
                        processed_data_to_store = df
                        print(f"  → Detected 'Week' column — directly extracted data for '{material}'")
                    elif has_date_column_overall:
                        # If a 'Date' column was found, try to apply H-based averaging
                        weekly_avg_df = calculate_weekly_averages(df)
                        
                        if "calculated_week" in [c.lower() for c in weekly_avg_df.columns]:
                            material_title = f"{material} (From file: {os.path.basename(file_path)} - Calculated from Date + H-blocks)"
                            weekly_avg_df.rename(columns={"Calculated_Week": "Week"}, inplace=True) # Rename for consistent output
                            processed_data_to_store = weekly_avg_df
                            print(f"  → Detected 'Date' column and 'H' markers in data columns — data processed for '{material}'")
                        else:
                            # If 'Date' but no 'H' for grouping, extract directly
                            material_title = f"{material} (From file: {os.path.basename(file_path)} - Directly Extracted - Date column present, no H-blocks in data columns)"
                            processed_data_to_store = df
                            print(f"  → Detected 'Date' column but no 'H' markers in data columns — data extracted directly for '{material}'")
                            
                    else: # No 'Week' or 'Date' column detected in overall headers
                        # Attempt H-based averaging as a primary method for grouping
                        weekly_avg_df = calculate_weekly_averages(df)
                        if "calculated_week" in [c.lower() for c in weekly_avg_df.columns]:
                            material_title = f"{material} (From file: {os.path.basename(file_path)} - Calculated from H in data columns)"
                            weekly_avg_df.rename(columns={"Calculated_Week": "Week"}, inplace=True)
                            processed_data_to_store = weekly_avg_df
                            print(f"  → No 'Date' or 'Week' column, but 'H' markers found in data columns — data processed for '{material}'")
                        else: 
                            # If no 'Week', 'Date', or 'H' markers found for grouping, extract directly
                            material_title = f"{material} (From file: {os.path.basename(file_path)} - Directly Extracted - No H, Date, or Week found for grouping)"
                            processed_data_to_store = df
                            print(f"  → No 'Date', 'Week', or 'H' markers found in data columns — data extracted directly for '{material}'")

                    # Store the processed data and its title for consolidated writing
                    if processed_data_to_store is not None and not processed_data_to_store.empty:
                        # Initialize list for this sheet if it doesn't exist yet
                        if sheet_name not in consolidated_output_data:
                            consolidated_output_data[sheet_name] = []
                        
                        # Append a tuple: (DataFrame for title, DataFrame for data)
                        # The title is a single-cell DataFrame for easy writing
                        title_df = pd.DataFrame({material: [material_title]})
                        consolidated_output_data[sheet_name].append((title_df, processed_data_to_store))

    # --- Final Step: Write all consolidated data to a single Excel file ---
    if not consolidated_output_data:
        print(f"\nNo data was processed from any file. Consolidated output file '{consolidated_output_file_name}' will not be created.")
        return

    print(f"\n--- Writing all processed data to consolidated file: {consolidated_output_file_name} ---")
    try:
        # Open the ExcelWriter once for the consolidated output file
        with pd.ExcelWriter(consolidated_output_file_name, engine='xlsxwriter') as writer:
            # Iterate through each sheet's collected data
            for sheet_name, list_of_df_pairs in consolidated_output_data.items():
                current_write_row = 0
                for title_df, data_df in list_of_df_pairs:
                    # Write the material title (e.g., "IS 2830... (From file: file1.xlsx)")
                    title_df.to_excel(writer, sheet_name=sheet_name, startrow=current_write_row, index=False, header=False)
                    current_write_row += len(title_df) # Move cursor past the title row

                    # Write the actual processed data
                    data_df.to_excel(writer, sheet_name=sheet_name, startrow=current_write_row, index=False)
                    current_write_row += len(data_df) + 2 # Move cursor past data and add 2 empty rows as buffer

                print(f"  Consolidated data written to sheet: '{sheet_name}'")
        print(f"✅ All consolidated processed data saved to '{consolidated_output_file_name}'")
    except Exception as e:
        print(f"An error occurred while writing the consolidated Excel file: {e}")

if __name__ == "__main__":
    main()